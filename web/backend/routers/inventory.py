from app.models.enums import StoreType
from app.services.product_service import ProductService
from decimal import Decimal
from fastapi import APIRouter, HTTPException, Query, File, UploadFile, Form
import csv
import io
from sqlalchemy import select
from sqlalchemy.orm import joinedload

from app.models.display_inventory import DisplayInventory
from app.models.inventory import Inventory
from app.models.enums import UserRole
from app.models.store import Store
from web.backend.dependencies import CurrentUser, SessionDep
from web.backend.schemas.stores import InventoryItemOut, ReceiveStockInput, BulkReceiveInput, DispatchDisplayInput
from app.bot.bot import bot
from app.services.notification_service import NotificationService
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
import logging
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/inventory", tags=["Inventory"])


@router.get(
    "/{store_id:int}",
    response_model=list[InventoryItemOut],
    summary="Остатки магазина / склада",
    description="Возвращает товары с количеством для указанного магазина (включая витрину).",
)
async def get_store_inventory(
    store_id: int,
    session: SessionDep,
    current_user: CurrentUser,
    include_empty: bool = Query(False),
) -> list[InventoryItemOut]:
    if current_user.role == UserRole.SELLER and current_user.store_id != store_id:
        raise HTTPException(status_code=403, detail="Access denied")

    # 1. Fetch regular inventory
    stmt_reg = (
        select(Inventory)
        .options(joinedload(Inventory.product))
        .where(Inventory.store_id == store_id)
    )
    if not include_empty:
        stmt_reg = stmt_reg.where(Inventory.quantity > 0)
    
    res_reg = await session.execute(stmt_reg)
    items_reg = res_reg.scalars().all()

    # 2. Fetch display inventory
    stmt_disp = (
        select(DisplayInventory)
        .options(joinedload(DisplayInventory.product))
        .where(DisplayInventory.store_id == store_id)
    )
    if not include_empty:
        stmt_disp = stmt_disp.where(DisplayInventory.quantity > 0)
    
    res_disp = await session.execute(stmt_disp)
    items_disp = res_disp.scalars().all()

    # Combine
    out = []
    for inv in items_reg:
        out.append(InventoryItemOut(
            product_id=inv.product_id,
            product_sku=inv.product.sku,
            quantity=inv.quantity,
            is_display=False
        ))
    
    for inv in items_disp:
        out.append(InventoryItemOut(
            product_id=inv.product_id,
            product_sku=inv.product.sku,
            quantity=inv.quantity,
            is_display=True
        ))

    return out


@router.get(
    "",
    response_model=dict,
    summary="Остатки всех магазинов",
    description="Возвращает остатки сгруппированные по магазинам (только Owner/Warehouse, включая витрину).",
)
async def get_all_inventory(
    session: SessionDep,
    current_user: CurrentUser,
) -> dict:
    if current_user.role not in (UserRole.OWNER, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Access denied")

    stores_result = await session.execute(select(Store).where(Store.is_active.is_(True)))
    stores = stores_result.scalars().all()
    store_map = {store.id: store.name for store in stores}
    result = {store.id: {"store_name": store.name, "items": []} for store in stores}

    # Regular inventory for all stores in one query.
    reg_result = await session.execute(
        select(Inventory)
        .options(joinedload(Inventory.product))
        .where(Inventory.quantity > 0, Inventory.store_id.in_(store_map.keys()))
    )
    for inv in reg_result.scalars().all():
        result[inv.store_id]["items"].append(
            {
                "product_id": inv.product_id,
                "sku": inv.product.sku,
                "quantity": inv.quantity,
                "is_display": False,
            }
        )

    # Display inventory for all stores in one query.
    disp_result = await session.execute(
        select(DisplayInventory)
        .options(joinedload(DisplayInventory.product))
        .where(DisplayInventory.quantity > 0, DisplayInventory.store_id.in_(store_map.keys()))
    )
    for inv in disp_result.scalars().all():
        result[inv.store_id]["items"].append(
            {
                "product_id": inv.product_id,
                "sku": inv.product.sku,
                "quantity": inv.quantity,
                "is_display": True,
            }
        )

    return result

@router.post(
    "/receive",
    response_model=dict,
    summary="Приход товара на Главный Склад",
    description="Добавляет количество существующего товара на Главный Склад (только Owner/Warehouse).",
)
async def receive_inventory(
    body: ReceiveStockInput,
    session: SessionDep,
    current_user: CurrentUser,
) -> dict:
    if current_user.role not in (UserRole.OWNER, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Access denied")

    from app.services.store_service import StoreService
    store_svc = StoreService(session)
    warehouse_id = await store_svc.get_main_warehouse_id()
    if not warehouse_id:
        raise HTTPException(status_code=404, detail="Главный склад не найден")

    from app.services.transaction_service import TransactionService
    txn_svc = TransactionService(session)
    
    try:
        inv = await txn_svc.receive_stock(
            warehouse_store_id=warehouse_id,
            product_id=body.product_id,
            quantity=body.quantity,
            user_id=current_user.id,
        )
        await session.commit()
    except ValueError as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    except Exception:
        await session.rollback()
        raise HTTPException(status_code=500, detail="Internal server error")
        
    return {
        "success": True, 
        "product_id": inv.product_id, 
        "new_quantity": inv.quantity
    }


@router.post(
    "/bulk-receive",
    response_model=dict,
    summary="Массовый приход товара из Excel",
    description="Принимает список SKU и количеств, обновляет остатки на Главном Складе (Создаёт новые товары если их нет).",
)
async def bulk_receive_inventory(
    body: BulkReceiveInput,
    session: SessionDep,
    current_user: CurrentUser,
) -> dict:
    if current_user.role not in (UserRole.OWNER, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Access denied")

    from app.services.store_service import StoreService
    from app.services.product_service import ProductService
    from app.services.transaction_service import TransactionService
    from app.models.product import Product
    
    store_svc = StoreService(session)
    product_svc = ProductService(session)
    txn_svc = TransactionService(session)
    
    warehouse_id = await store_svc.get_main_warehouse_id()
    if not warehouse_id:
        raise HTTPException(status_code=404, detail="Главный склад не найден")

    updated_count = 0
    created_count = 0
    
    try:
        for item in body.items:
            sku = item.sku.strip().upper()
            if not sku:
                raise HTTPException(status_code=400, detail="SKU cannot be empty")
            if item.quantity <= 0:
                raise HTTPException(status_code=400, detail=f"Invalid quantity for SKU {sku}")

            # Find product by SKU
            stmt = select(Product).where(Product.sku == sku)
            res = await session.execute(stmt)
            product = res.scalar_one_or_none()
            
            # Resolve effective brand: per-item brand is required
            effective_brand: str | None = (
                item.brand.strip() if item.brand and item.brand.strip()
                else (body.default_brand.strip() if body.default_brand else None)
            )
            if not effective_brand:
                raise HTTPException(
                    status_code=400,
                    detail=f"Фирма обязательна для товара SKU='{sku}'. Выберите фирму для каждого товара."
                )

            if not product:
                # Create new product
                price_val = Decimal(str(item.price)) if item.price is not None else Decimal(0)
                if price_val < 0:
                    raise HTTPException(status_code=400, detail=f"Invalid price for SKU {sku}")
                product = await product_svc.create_product(
                    sku=sku,
                    price=price_val,
                    brand=effective_brand,
                )
                created_count += 1
            else:
                # Update existing product — apply price and brand if provided
                if item.price is not None:
                    new_price = Decimal(str(item.price))
                    if new_price < 0:
                        raise HTTPException(status_code=400, detail=f"Invalid price for SKU {sku}")
                    product.price = new_price
                if effective_brand:
                    product.brand = effective_brand
                
            # Receive stock — replace or accumulate based on flag
            inv_stmt = select(Inventory).where(
                Inventory.store_id == warehouse_id,
                Inventory.product_id == product.id,
            )
            inv_res = await session.execute(inv_stmt)
            existing_inv = inv_res.scalar_one_or_none()

            if existing_inv:
                if body.replace_quantity:
                    existing_inv.quantity = item.quantity
                else:
                    existing_inv.quantity += item.quantity
            else:
                session.add(Inventory(
                    store_id=warehouse_id,
                    product_id=product.id,
                    quantity=item.quantity,
                ))
            updated_count += 1
            
        await session.commit()
    except HTTPException:
        await session.rollback()
        raise
    except Exception as e:
        await session.rollback()
        logger.exception("bulk_receive_inventory error: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Ошибка базы данных: {str(e)}")
    
    return {
        "success": True,
        "processed": updated_count,
        "created": created_count,
    }


@router.post(
    "/dispatch-display",
    response_model=dict,
    summary="Отправить образцы (витрину) в магазин",
    description="Перемещает товар с Главного Склада в статус 'Образец' для магазина.",
)
async def dispatch_display(
    body: DispatchDisplayInput,
    session: SessionDep,
    current_user: CurrentUser,
) -> dict:
    if current_user.role not in (UserRole.OWNER, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Access denied")

    from app.services.store_service import StoreService
    store_svc = StoreService(session)
    warehouse_id = await store_svc.get_main_warehouse_id()
    if not warehouse_id:
        raise HTTPException(status_code=404, detail="Главный склад не найден")

    from app.services.transaction_service import TransactionService
    txn_svc = TransactionService(session)
    
    try:
        display_order, wh_inv = await txn_svc.dispatch_display_items(
            warehouse_store_id=warehouse_id,
            target_store_id=body.target_store_id,
            product_id=body.product_id,
            quantity=body.quantity,
            user_id=current_user.id,
        )
        await session.commit()
    except ValueError as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    except Exception:
        await session.rollback()
        raise HTTPException(status_code=500, detail="Internal server error")
    
    # Send notification to sellers
    try:
        from app.models.product import Product
        res = await session.execute(select(Product).where(Product.id == display_order.product_id))
        product = res.scalar_one()
        
        notif_svc = NotificationService(bot, session)
        await notif_svc.notify_sellers(
            store_id=display_order.store_id,
            text=lambda t: t("display_dispatched_seller_notif", sku=product.sku, qty=display_order.quantity),
            reply_markup=lambda t: InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text=t("btn_received"), 
                        callback_data=f"display:receive:{display_order.id}"
                    ),
                    InlineKeyboardButton(
                        text=t("btn_not_received"), 
                        callback_data=f"display:reject:{display_order.id}"
                    )
                ]
            ])
        )
    except Exception as exc:
        logger.error("Failed to send web dispatch notification: %s", exc)
        
    return {
        "success": True, 
        "order_id": display_order.id,
        "product_id": display_order.product_id, 
        "target_store_id": display_order.store_id,
        "quantity": display_order.quantity
    }


@router.post(
    "/import-csv",
    response_model=dict,
    summary="Импорт товаров из CSV файла",
    description="Загружает CSV файл, создает товары и добавляет остатки в указанный магазин.",
)
async def import_csv_endpoint(
    session: SessionDep,
    current_user: CurrentUser,
    file: UploadFile = File(...),
    store_id: int = Form(None)
) -> dict:
    if current_user.role not in (UserRole.OWNER, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Access denied")

    from app.models.product import Product
    from app.services.store_service import StoreService
    
    store_svc = StoreService(session)
    
    if store_id:
        store = await session.get(Store, store_id)
        if not store:
            raise HTTPException(status_code=404, detail="Указанный магазин/склад не найден")
    else:
        warehouse_id = await store_svc.get_main_warehouse_id()
        if not warehouse_id:
            raise HTTPException(status_code=404, detail="Главный склад не найден")
        store = await session.get(Store, warehouse_id)

    try:
        content = await file.read()
        
        headers = []
        rows = []
        
        if file.filename and file.filename.lower().endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            # Read all sheets searching for a header
            header_row_idx = -1
            active_headers = []
            active_rows = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_data = list(sheet.iter_rows(values_only=True))
                if not sheet_data:
                    continue
                    
                for i, row_data in enumerate(sheet_data):
                    # Check if this row looks like a header (Sku, Артикул, Название, Наим, Кол, Цена, S/N)
                    r_str = [str(cell).strip().lower() if cell is not None else "" for cell in row_data]
                    keywords = ["sku", "артикул", "название", "наим", "кол", "цена", "price", "s/n", "serial"]
                    
                    if any(k in h for h in r_str for k in keywords):
                        header_row_idx = i
                        active_headers = r_str
                        # Found it! Now collect the rest of the rows in this sheet
                        for rd in sheet_data[i+1:]:
                            if any(rd): # Skip totally empty rows
                                active_rows.append([str(c).strip() if c is not None else "" for c in rd])
                        break
                
                if header_row_idx != -1:
                    headers = active_headers
                    rows = active_rows
                    logger.info("Found header in sheet '%s' at row %d: %s", sheet_name, header_row_idx, headers)
                    break
            
            if header_row_idx == -1:
                first_sheet = wb.sheetnames[0]
                sheet_0 = wb[first_sheet]
                rows_0 = list(sheet_0.iter_rows(values_only=True))
                first_row = [str(cell) for cell in rows_0[0]] if rows_0 else "empty"
                raise HTTPException(status_code=400, detail=f"Не удалось найти таблицу. Проверьте названия колонок (Sku, Кол-во, Цена). Первая строка листа '{first_sheet}': {first_row}")
        else:
            # Decode considering BOM for Excel CSV exports
            try:
                decoded_content = content.decode('utf-8-sig')
            except UnicodeDecodeError:
                try:
                    decoded_content = content.decode('windows-1251')
                except Exception:
                    raise HTTPException(status_code=400, detail="Не удалось определить кодировку файла (попробуйте UTF-8 или Excel)")
                
            all_lines = [
                line.strip() for line in decoded_content.splitlines() if line.strip()
            ]
            
            # Smart delimiter detection
            delimiter = ";"
            if all_lines:
                # Try common delimiters and see which gives more columns in the first line
                comma_cols = len(all_lines[0].split(","))
                semicolon_cols = len(all_lines[0].split(";"))
                if comma_cols > semicolon_cols:
                    delimiter = ","
                    logger.info("Detected comma delimiter (cols: %d)", comma_cols)
            
            reader = csv.reader(all_lines, delimiter=delimiter)
            
            header_found = False
            for row_data in reader:
                if not row_data:
                    continue
                row_str = [h.strip().lower() for h in row_data]
                if not header_found:
                    # Check for keywords
                    keywords = ["sku", "артикул", "название", "наим", "кол", "цена", "price", "s/n", "serial"]
                    if any(k in h for h in row_str for k in keywords):
                        headers = row_str
                        header_found = True
                        logger.info("Found CSV header: %s", headers)
                    continue
                rows.append(row_data)
                
            if not header_found:
                first_line = all_lines[0] if all_lines else "empty"
                raise HTTPException(status_code=400, detail=f"Не найдена шапка в CSV. Первая строка: {first_line}. Проверьте заголовки: Sku, Кол-во, Цена")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Import error")
        raise HTTPException(status_code=400, detail=f"Ошибка чтения файла: {str(e)}")

    idx_sku, idx_qty, idx_price = -1, -1, -1
    
    # Smarter mapping: check for partial matches and exact matches
    for i, h in enumerate(headers):
        h_clean = h.strip().lower()
        if h_clean in ["sku", "артикул", "код", "товар"]:
            idx_sku = i
        elif "sku" in h_clean or "артикул" in h_clean or "название" in h_clean:
            if idx_sku == -1:
                idx_sku = i
            
        if any(x in h_clean for x in ["kolichestva", "kolichestvo", "склад", "количество", "qty", "шт", "остаток"]):
            idx_qty = i
        elif any(x in h_clean for x in ["sena", "цена", "price", "cost", "сумма", "стоимость"]):
            idx_price = i

    if idx_sku == -1:
        raise HTTPException(status_code=400, detail=f"В шапке таблицы не найдена колонка для SKU (артикула). Найденные колонки: {headers}")

    def parse_qty(val: str):
        try:
            return int(float(val)) if val else 0
        except Exception:
            return 0
            
    def parse_price(val: str):
        try:
            val = val.replace(',', '.')
            return Decimal(val) if val else Decimal("0.00")
        except Exception:
            return Decimal("0.00")

    products_to_add = 0
    products_updated = 0
    qty_added = 0

    try:
        parsed_rows = []
        skus: set[str] = set()
        for row in rows:
            if not row or len(row) <= idx_sku:
                continue
            
            sku = str(row[idx_sku]).strip()
            if not sku or "пример" in sku.lower() or sku.lower() == "sku":
                continue
            
            qty = parse_qty(str(row[idx_qty])) if idx_qty != -1 and len(row) > idx_qty else 0
            price = parse_price(str(row[idx_price])) if idx_price != -1 and len(row) > idx_price else Decimal("0.00")
            parsed_rows.append({"sku": sku, "qty": qty, "price": price})
            skus.add(sku)

        products_result = await session.execute(select(Product).where(Product.sku.in_(skus)))
        products_map: dict[str, Product] = {
            product.sku: product for product in products_result.scalars().all()
        }

        for row in parsed_rows:
            sku = row["sku"]
            price = row["price"]
            product = products_map.get(sku)
            if not product:
                product = Product(
                    sku=sku,
                    brand=ProductService.infer_brand_from_sku(sku),
                    price=price,
                    is_active=True,
                )
                session.add(product)
                products_map[sku] = product
                products_to_add += 1
            else:
                if price > 0:
                    product.price = price
                products_updated += 1

        await session.flush()

        inventory_model = DisplayInventory if store.store_type == StoreType.STORE else Inventory
        product_ids = list({
            product.id for row in parsed_rows
            if row["qty"] > 0
            for product in [products_map[row["sku"]]]
            if product.id is not None
        })
        inventory_map: dict[int, DisplayInventory | Inventory] = {}
        if product_ids:
            inventory_result = await session.execute(
                select(inventory_model).where(
                    inventory_model.store_id == store.id,
                    inventory_model.product_id.in_(product_ids),
                )
            )
            inventory_map = {
                inventory.product_id: inventory
                for inventory in inventory_result.scalars().all()
            }

        for row in parsed_rows:
            qty = row["qty"]
            if qty <= 0:
                continue

            product = products_map[row["sku"]]
            inventory = inventory_map.get(product.id)

            if not inventory:
                inventory = inventory_model(
                    store_id=store.id,
                    product_id=product.id,
                    quantity=qty,
                )
                session.add(inventory)
                inventory_map[product.id] = inventory
            else:
                inventory.quantity += qty
            qty_added += qty

        await session.commit()
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка базы данных: {str(e)}")

    return {
        "success": True,
        "store": store.name,
        "created": products_to_add,
        "updated": products_updated,
        "added_qty": qty_added
    }
